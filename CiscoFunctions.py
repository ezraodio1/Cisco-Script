#import all necessary modules and methods
import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
import parser
import dateparser
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.styles.borders import Border, Side

#The function below sets up variables and the spreadsheet. Needs to be first function called.
def setup(file, newSheet):
    #setup variables
    global filename
    global workbook
    global original_sheet
    global modified_sheet
    global numRows
    global numCols
    global count
    global EUCost
    count = 0
    
    #Choose sheet to work on and change title
    filename = file
    workbook = openpyxl.load_workbook(filename)
    original_sheet = workbook.active
    original_sheet.title = "Cisco Item Details"

    #get number of rows and number of columns in spreadsheet
    numRows = original_sheet.max_row
    numCols = original_sheet.max_column
    
    #The loop below gets the row number of "Quote Name"
    for row in original_sheet.iter_rows(min_row = 1, max_col = numCols, max_row = numRows): 
        for cell in row:
            if cell.value == "Quote Name":
                quoteNameRow = cell.row
                break

    #The loop below gets the row number of Quote Details
    for row in original_sheet.iter_rows(min_row = 1, max_col = numCols, max_row = numRows): #get row number of "Quote Details"
        for cell in row:
            if cell.value == "Quote Details":
                quoteDetailsRow = cell.row
                break

    #The two lines below delete the unnecessary rows between "Quote Name" and "Quote Details" at the top of the spreadsheet 
    rowsToDelete = quoteDetailsRow - quoteNameRow + 1
    original_sheet.delete_rows(quoteNameRow, rowsToDelete)

    '''for row in original_sheet.iter_rows(min_row = 1, max_col = numCols, max_row = numRows): #get row number of "Quote Details"
        for cell in row:
            if cell.value == "Date":
                quoteDateRow = cell.row
                break

    original_sheet.delete_rows(0, quoteDateRow)'''

    #The loop below creates a new sheet that is a copy of the original sheet
    modified_sheet = workbook.create_sheet(newSheet, index = 0)
    for row in original_sheet: 
        for cell in row:
            modified_sheet[cell.coordinate].value = cell.value

    #Make the copy sheet the one we are working on
    workbook.active = modified_sheet
    
    numRows = modified_sheet.max_row
    numCols = modified_sheet.max_column

def addFilters():
    for row in original_sheet.iter_rows(min_row = 1, max_col = numCols, max_row = numRows): #get row number of "Quote Name"
        for cell in row:
            if cell.value == "Product Number":
                filterRow = cell.row
                break
    
    FullRange = "A" + str(filterRow) + ":" + get_column_letter(original_sheet.max_column) + str(original_sheet.max_row)
    original_sheet.auto_filter.ref = FullRange

    maxRowMod = modified_sheet.max_row - count
    maxColMod = modified_sheet.max_column - count
    
    FullRange2 = "A" + str(filterRow) + ":" + get_column_letter(maxColMod) + str(maxRowMod)
    modified_sheet.auto_filter.ref = FullRange2

def calcEUCost():
    global EUCost
    EUCost = 0
    DOL = float(input("Enter CCS's discount off list price. Ex: if discount off list is 20%, enter '20' "))
    margin = float(input("Enter your desired margin. Ex: if margin is 20%, enter '20' "))
    numCols = original_sheet.max_column
    numRows = original_sheet.max_row
    for row in modified_sheet.iter_rows(min_row = 1, max_col = numCols, max_row = numRows):
        for cell in row:
            if cell.value == "Discount":
                startRow = cell.row + 1
                costColumn = cell.column + 1
                modified_sheet.insert_cols(costColumn)
                modified_sheet.cell(startRow - 1, costColumn).value = "End Customer Cost"
                break

    for row in modified_sheet.iter_rows(min_row = 1, max_col = numCols, max_row = numRows):
        for cell in row:
            if cell.value == "Extended List Price" or cell.value == " Extended List Price":
                listCol = cell.column
                break

    for column in modified_sheet.iter_cols(min_col = costColumn, max_row = numRows, max_col = costColumn):
        for cell in column:
            if cell.row >= startRow:
                currentRow = cell.row
                cell.value = round(float(modified_sheet.cell(currentRow, listCol).value)*(100 - DOL)/(100-margin), 2)


    for row in original_sheet.iter_rows(min_row = 1, max_col = numCols, max_row = numRows):
        for cell in row:
            if cell.value == "Discount":
                startRow = cell.row + 1
                costColumn = cell.column + 1
                original_sheet.insert_cols(costColumn)
                original_sheet.cell(startRow - 1, costColumn).value = "End Customer Cost"
                break

    for row in original_sheet.iter_rows(min_row = 1, max_col = numCols, max_row = numRows):
        for cell in row:
            if cell.value == "Extended List Price" or cell.value == " Extended List Price":
                listCol = cell.column
                break

    for column in original_sheet.iter_cols(min_col = costColumn, max_row = numRows, max_col = costColumn):
        for cell in column:
            if cell.row >= startRow:
                currentRow = cell.row
                cell.value = round(float(original_sheet.cell(currentRow, listCol).value)*(100 - DOL)/(100-margin), 2)
                EUCost = EUCost + cell.value
                

def calcDiscount():
    for row in modified_sheet.iter_rows(min_row = 1, max_col = numCols, max_row = numRows): #get row number of "Quote Name"
        for cell in row:
            if cell.value == "Discount":
                startRow = cell.row + 1
                discCol = cell.column
                break

    for row in modified_sheet.iter_rows(min_row = 1, max_col = numCols, max_row = numRows): #get row number of "Quote Name"
        for cell in row:
            if cell.value == "Extended List Price" or cell.value == " Extended List Price":
                extListCol = cell.column
                break

    for row in modified_sheet.iter_rows(min_row = 1, max_col = numCols, max_row = numRows): #get row number of "Quote Name"
        for cell in row:
            if cell.value == "End Customer Cost":
                custCostCol = cell.column
                break

    for column in modified_sheet.iter_cols(min_col = discCol, max_row = numRows, max_col = discCol):
        for cell in column:
            if cell.row >= startRow:
                currentRow = cell.row
                cost = float(modified_sheet.cell(currentRow, custCostCol).value)
                listPrice = float(modified_sheet.cell(currentRow, extListCol).value)

                if listPrice > 0:
                    cell.value = round(100 * (1 - cost/listPrice), 1)
                else:
                    cell.value = 0
                #cell.value = 100*(1 - float(modified_sheet.cell(currentRow, custCostCol).value) / float(modified_sheet.cell(currentRow, extListCol).value))

    
    
def deleteCol(name):
    for row in modified_sheet.iter_rows(min_row = 1, max_col = numCols, max_row = numRows):
        for cell in row:
            if cell.value == name:
                modified_sheet.delete_cols(cell.column)
                break
    global count
    count = count + 1

def hideRows():
    numRows = original_sheet.max_row
    for row in modified_sheet.iter_rows(min_row = 1, max_col = numCols, max_row = numRows):
        for cell in row:
            if cell.value == "Parent Instance Number":
                parentInstanceRow = cell.row
                parentInstanceColumn = cell.column
                break

    parentNum = modified_sheet.cell(parentInstanceRow + 1, parentInstanceColumn).value #get first Parent Instance Number
    numToHide = 0

    startHide = parentInstanceRow + 1
    endHide = parentInstanceRow + 1

    for i in range(parentInstanceRow + 1, numRows + 1):
        if modified_sheet.cell(i, parentInstanceColumn).value == parentNum:
            endHide = i

            if i == numRows:
                for idx in range(startHide + 1, endHide + 1):
                    modified_sheet.row_dimensions[idx].hidden = True
                    
        else:
            if endHide != startHide:
                for idx in range(startHide + 1, endHide + 1):
                    modified_sheet.row_dimensions[idx].hidden = True

            parentNum = modified_sheet.cell(i, parentInstanceColumn).value
            startHide = i
            endHide = i

            

    
    '''for i in range(parentInstanceRow + 2, numRows):
        if (modified_sheet.cell(i, parentInstanceColumn).value == parentNum):
            numToHide = numToHide + 1

            if i == numRows:
                modified_sheet.row_dimensions.group(i - numToHide + 1, i + 1, hidden = True)
        
        else:
            if numToHide > 0:
                modified_sheet.row_dimensions.group(i - numToHide - 1, i, hidden=True)
                numToHide = 0

            parentNum = modified_sheet.cell(i, parentInstanceColumn).value'''    
    

def highlight():

    Color = True

    blueFill = PatternFill(start_color='3e55c1',
                       end_color='3e55c1',
                       fill_type='solid')

    for row in original_sheet.iter_rows(min_row = 1, max_col = original_sheet.max_column, max_row = original_sheet.max_row): #get column and row number of "Major/Minor"
        for cell in row:
            if cell.value == "Parent Instance Number":
                OriginalParentRow = cell.row
                OriginalParentColumn = cell.column

    OriginalParentNum = original_sheet.cell(OriginalParentRow + 1, OriginalParentColumn).value

    for i in range(OriginalParentRow + 1, original_sheet.max_row + 1):
        if (original_sheet.cell(i, OriginalParentColumn).value == OriginalParentNum):
            if Color == True:
                for n in range(1, original_sheet.max_column + 1):
                    original_sheet.cell(i, n).fill = blueFill
        else:
            Color = not Color
            OriginalParentNum = original_sheet.cell(i, OriginalParentColumn).value
            if Color == True:
                for n in range(1, original_sheet.max_column + 1):
                    original_sheet.cell(i, n).fill = blueFill

def warnClient():
    numCols = original_sheet.max_column
    numRows = original_sheet.max_row
    redFill = PatternFill(start_color='c47f12',
                       end_color='c47f12',
                       fill_type='solid')
    
    for row in original_sheet.iter_rows(min_row = 1, max_col = numCols, max_row = numRows): #get row number of "Last Date of Support"
        for cell in row:
            if cell.value == "Last Date of Support":
                EOSL_Row = cell.row + 1
                EOSL_Column = cell.column
                
            if cell.value == "End Date":
                endDateCol = cell.column

    for column in original_sheet.iter_cols(min_col = EOSL_Column, max_row = numRows, max_col = EOSL_Column):
        for cell in column:
            if cell.value != '' and cell.value is not None and cell.value != "Last Date of Support":
                sameRow = cell.row
                date1 = dateparser.parse(cell.value)
                date2 = dateparser.parse(original_sheet.cell(sameRow, endDateCol).value)
                if date1 is not None:
                    if date1 <= date2:
                        cell.fill = redFill
                        original_sheet.cell(sameRow, endDateCol).fill = redFill

    numCols = modified_sheet.max_column
    numRows = modified_sheet.max_row
    
    for row in modified_sheet.iter_rows(min_row = 1, max_col = numCols, max_row = numRows): #get row number of "Last Date of Support"
        for cell in row:
            if cell.value == "Last Date of Support":
                EOSL_Row = cell.row + 1
                EOSL_Column = cell.column
                
            if cell.value == "End Date":
                endDateCol = cell.column

    '''redFill = PatternFill(start_color='ff6e6e',
                    end_color='ff6e6e',
                    fill_type='solid')'''
        
    for column in modified_sheet.iter_cols(min_col = EOSL_Column, max_row = numRows, max_col = EOSL_Column):
        for cell in column:
            if cell.value != '' and cell.value is not None and cell.value != "Last Date of Support":
                sameRow = cell.row
                date1 = dateparser.parse(cell.value)
                date2 = dateparser.parse(modified_sheet.cell(sameRow, endDateCol).value)
                if date1 is not None:
                    if date1 <= date2:
                        cell.fill = redFill
                        i = cell.row
                        for n in range(1, modified_sheet.max_column - count):
                            modified_sheet.cell(i, n).fill = redFill
    
                

def resizeColumns():
    column_widths = []
    for row in modified_sheet.iter_rows():
        for i, cell in enumerate(row):
            try:
                column_widths[i] = max(column_widths[i], len(str(cell.value)))
            except IndexError:
                column_widths.append(len(str(cell.value)))

    for i, column_width in enumerate(column_widths):
        modified_sheet.column_dimensions[get_column_letter(i + 1)].width = column_width

    column_widths2 = []
    for row in original_sheet.iter_rows():
        for i, cell in enumerate(row):
            try:
                column_widths2[i] = max(column_widths2[i], len(str(cell.value)))
            except IndexError:
                column_widths2.append(len(str(cell.value)))

    for i, column_width in enumerate(column_widths2):
        original_sheet.column_dimensions[get_column_letter(i + 1)].width = column_width

    
    #for n in range(1, modified_sheet.max_column):
        #modified_sheet.column_dimensions[n].bestFit = True
        #modified_sheet.column_dimensions[n] = ColumnDimension(modified_sheet, bestFit = True)
        #ColumnDimension(modified_sheet, bestFit = True)

    #for n in range(1, original_sheet.max_column):
        #original_sheet.column_dimensions[n].bestFit = True
        #original_sheet.dimensions.ColumnDimension(bestFit = True)
    
def makePretty():
    colorBorder = Border(bottom=Side(border_style = 'thick', color = '7da40e'))
    
    for row in modified_sheet.iter_rows(min_row = 1, max_col = numCols, max_row = numRows):
        for cell in row:
            if cell.value == "Product Number":
                boldRow = cell.row
                break
    for n in range(1, numCols - count + 2):
        modified_sheet.cell(boldRow, n).font = Font(bold = True)
        modified_sheet.cell(boldRow, n).border = colorBorder
    for n in range(1, numCols + 2):
        original_sheet.cell(boldRow, n).font = Font(bold = True)
        original_sheet.cell(boldRow, n).border = colorBorder
    #add bold and underline


def deleteRows():

    for row in original_sheet.iter_rows(min_row = 1, max_col = numCols, max_row = numRows):
        for cell in row:
            if cell.value == "Date":
                dateRow = cell.row
                break
    original_sheet.delete_rows(1, dateRow - 1)

    '''for row in modified_sheet.iter_rows(min_row = 1, max_col = numCols, max_row = numRows):
        for cell in row:
            if cell.value == "Date":
                dateRow = cell.row
                break'''
    modified_sheet.delete_rows(1, dateRow - 1)
    
def name():
    skuLength = 0
    descLength = 0

    for row in modified_sheet.iter_rows(min_row = 1, max_col = numCols, max_row = numRows):
        for cell in row:
            if cell.value == "SKU":
                skuCol = cell.column
                skuRow = cell.row
                break
    for n in range (skuRow + 1, numRows):
        try:
            skuLength = skuLength + len(modified_sheet.cell(n, skuCol).value)

        except Exception:
            pass

    for row in modified_sheet.iter_rows(min_row = 1, max_col = numCols, max_row = numRows):
        for cell in row:
            if cell.value == "Product Description":
                descCol = cell.column
                descRow = cell.row
                break
            
    for n in range (descRow + 1, numRows):
        try:
            descLength = descLength + len(modified_sheet.cell(n, descCol).value)

        except Exception:
            pass

    seedNum = numRows
    seedNum = seedNum * 7177
    seedNum = seedNum + EUCost
    seedNum = seedNum * 7177
    seedNum = seedNum + skuLength
    seedNum = seedNum * 7177
    seedNum = seedNum + descLength

    seedNum = int(seedNum)
    hexSeed = hex(seedNum)

    hexSeed = str(hexSeed)
    num1 = int(hexSeed[len(hexSeed) - 1:], 16)
    num2 = int(hexSeed[len(hexSeed) - 2:-1], 16)
    num3 = int(hexSeed[len(hexSeed) - 3:-2], 16)
    num4 = int(hexSeed[len(hexSeed) - 4:-3], 16)

    num1 = bin(num1)[2:]
    num2 = bin(num2)[2:]
    num3 = bin(num3)[2:]
    num4 = bin(num4)[2:]


    num1 = int(num1 + num4[:1], 2)
    num2 = int(num2 + num4[1:2], 2)
    num3 = int(num3 + num4[2:3], 2)

    ref = 'ABCDEFGHJKLMNPQRSTUVWXYZ23456789'

    final = ref[num1 - 1:num1] + ref[num2 - 1:num2] + ref[num3 - 1:num3]
    if final == 'ASS':
        final = 'AST'
    if final == 'KKK':
        final = 'KKL'
    if final == 'WTF':
        final = 'WTG'


    clientName = input('Please enter the client name ')
    newFilename = clientName + '_CiscoRenewalDetails_DocID:' + final + '_' + datetime.now().strftime('%m-%d-%y') + '.xlsx'

    print(newFilename)
    workbook.save(newFilename)

    
